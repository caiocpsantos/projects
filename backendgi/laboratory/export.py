from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import HttpResponse
import csv
from .models import LabTestWater, LabTestIndustrialSoil, LabTestResidentialSoil, LabTestAgriculturalSoil, LabTestIndustrialSteam, LabTestResidentialSteam
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter




class ExportLabTestWaterCSV(APIView):
    """
    API para exportar dados de LabTestWater como CSV.
    O nome do arquivo inclui report_name e sample_type.
    """

    def get(self, request):
        report_name = request.data.get("report_name")   
        sample_type = request.data.get("sample_type") 
        # Identificar todas as sample_names únicas dinamicamente
        
        queryset = LabTestWater.objects.all()
        sample_names = queryset.values_list('sample_name', flat=True).distinct()

        # Criar a resposta HTTP para o arquivo CSV
        file_name = f"{report_name}_{sample_type}.csv"
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        # Criar o writer do CSV
        writer = csv.writer(response)

        # Escrever o cabeçalho
        header = ['substance', 'agency', 'intervention_value', 'measure'] + list(sample_names)
        writer.writerow(header)

        # Agrupar os dados por substance
        substances = queryset.values('substance', 'agency', 'intervention_value', 'measure').distinct()

        for substance_data in substances:
            # Linha base com substance, agency e intervention_value
            row = [
                substance_data['substance'],
                substance_data['agency'],
                substance_data['intervention_value'],
                substance_data['measure']
            ]

            # Adicionar os resultados para cada sample_name
            results = []
            for sample_name in sample_names:
                result = queryset.filter(
                    substance=substance_data['substance'],
                    sample_name=sample_name
                ).first()

                # Adicionar o result ou vazio se não houver dados
                if result:
                    if result.result > 0:
                        results.append(result.result)
                    else: results.append(result.under_limit)
                else: results.append('Não Testado')

            # Completar a linha e escrever no CSV
            writer.writerow(row + results)

        return response
    


class ExportLabTestWaterXLSX(APIView):
    """
    API para exportar dados de LabTestWater como XLSX.
    O nome do arquivo inclui report_name e sample_type.
    """

    def get(self, request):
        report_name = request.data.get("report_name")   
        sample_type = request.data.get("sample_type") 

        # Identificar todas as sample_names únicas dinamicamente
        queryset = LabTestWater.objects.all()
        sample_names = queryset.values_list('sample_name', flat=True).distinct()

        # Criar o arquivo Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "LabTestWater Data"

        # Escrever o cabeçalho
        header = ['substance', 'agency', 'intervention_value', 'measure'] + list(sample_names)
        for col_num, column_title in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = column_title
            sheet[f"{col_letter}1"].font = Font(bold=True)

        # Agrupar os dados por substance
        substances = queryset.values('substance', 'agency', 'intervention_value', 'measure').distinct()

        row_number = 2  # Começa na segunda linha (primeira é o cabeçalho)
        for substance_data in substances:
            # Linha base com substance, agency e intervention_value
            row = [
                substance_data['substance'],
                substance_data['agency'],
                substance_data['intervention_value'],
                substance_data['measure']
            ]

            # Adicionar os resultados para cada sample_name
            results = []
            for sample_name in sample_names:
                result = queryset.filter(
                    substance=substance_data['substance'],
                    sample_name=sample_name
                ).first()

                if result:
                    if result.result > 0:
                        results.append(result.result)
                    else:
                        results.append(result.under_limit)
                else:
                    results.append('Não Testado')

            # Completar a linha com os dados
            full_row = row + results

            for col_num, cell_value in enumerate(full_row, 1):
                cell = sheet.cell(row=row_number, column=col_num, value=cell_value)

                # Aplicar a formatação condicional apenas para os valores de result
                if col_num > 4:  # Resultados começam na 5ª coluna
                    if isinstance(cell_value,float):    
                        intervention_value = substance_data['intervention_value']
                        if intervention_value is not None:
                            if cell_value > intervention_value:
                                cell.font = Font(color="FF0000")  # Texto branco para contraste
                    else:
                        cell.font = Font(color="8a8a8a")            

            row_number += 1

        # Criar a resposta HTTP para o arquivo XLSX
        file_name = f"{report_name}_{sample_type}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        workbook.save(response)
        return response    
    

######################################### Solo Industrial #######################################################################



class ExportLabTestIndustrialSoilCSV(APIView):
    """
    API para exportar dados de LabTestWater como CSV.
    O nome do arquivo inclui report_name e sample_type.
    """

    def get(self, request):
        report_name = request.data.get("report_name")   
        sample_type = request.data.get("sample_type") 
        # Identificar todas as sample_names únicas dinamicamente
        
        queryset = LabTestIndustrialSoil.objects.all()
        sample_names = queryset.values_list('sample_name', flat=True).distinct()

        # Criar a resposta HTTP para o arquivo CSV
        file_name = f"{report_name}_{sample_type}.csv"
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        # Criar o writer do CSV
        writer = csv.writer(response)

        # Escrever o cabeçalho
        header = ['substance', 'agency', 'intervention_value', 'measure'] + list(sample_names)
        writer.writerow(header)

        # Agrupar os dados por substance
        substances = queryset.values('substance', 'agency', 'intervention_value', 'measure').distinct()

        for substance_data in substances:
            # Linha base com substance, agency e intervention_value
            row = [
                substance_data['substance'],
                substance_data['agency'],
                substance_data['intervention_value'],
                substance_data['measure']
            ]

            # Adicionar os resultados para cada sample_name
            results = []
            for sample_name in sample_names:
                result = queryset.filter(
                    substance=substance_data['substance'],
                    sample_name=sample_name
                ).first()

                # Adicionar o result ou vazio se não houver dados
                if result:
                    if result.result > 0:
                        results.append(result.result)
                    else: results.append(result.under_limit)
                else: results.append('Não Testado')

            # Completar a linha e escrever no CSV
            writer.writerow(row + results)

        return response
    


class ExportLabTestIndustrialSoilXLSX(APIView):
    """
    API para exportar dados de LabTestWater como XLSX.
    O nome do arquivo inclui report_name e sample_type.
    """

    def get(self, request):
        report_name = request.data.get("report_name")   
        sample_type = request.data.get("sample_type") 

        # Identificar todas as sample_names únicas dinamicamente
        queryset = LabTestIndustrialSoil.objects.all()
        sample_names = queryset.values_list('sample_name', flat=True).distinct()

        # Criar o arquivo Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "LabTestWater Data"

        # Escrever o cabeçalho
        header = ['substance', 'agency', 'intervention_value', 'measure'] + list(sample_names)
        for col_num, column_title in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = column_title
            sheet[f"{col_letter}1"].font = Font(bold=True)

        # Agrupar os dados por substance
        substances = queryset.values('substance', 'agency', 'intervention_value', 'measure').distinct()

        row_number = 2  # Começa na segunda linha (primeira é o cabeçalho)
        for substance_data in substances:
            # Linha base com substance, agency e intervention_value
            row = [
                substance_data['substance'],
                substance_data['agency'],
                substance_data['intervention_value'],
                substance_data['measure']
            ]

            # Adicionar os resultados para cada sample_name
            results = []
            for sample_name in sample_names:
                result = queryset.filter(
                    substance=substance_data['substance'],
                    sample_name=sample_name
                ).first()

                if result:
                    if result.result > 0:
                        results.append(result.result)
                    else:
                        results.append(result.under_limit)
                else:
                    results.append('Não Testado')

            # Completar a linha com os dados
            full_row = row + results

            for col_num, cell_value in enumerate(full_row, 1):
                cell = sheet.cell(row=row_number, column=col_num, value=cell_value)

                # Aplicar a formatação condicional apenas para os valores de result
                if col_num > 4:  # Resultados começam na 5ª coluna
                    if isinstance(cell_value,float):    
                        intervention_value = substance_data['intervention_value']
                        if intervention_value is not None:
                            if cell_value > intervention_value:
                                cell.font = Font(color="FF0000") 
                    else:
                        cell.font = Font(color="8a8a8a")            

            row_number += 1

        # Criar a resposta HTTP para o arquivo XLSX
        file_name = f"{report_name}_{sample_type}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        workbook.save(response)
        return response

######################################### Solo Agrícola #######################################################################



class ExportLabTestAgriculturalSoilCSV(APIView):
    """
    API para exportar dados de LabTestWater como CSV.
    O nome do arquivo inclui report_name e sample_type.
    """

    def get(self, request):
        report_name = request.data.get("report_name")   
        sample_type = request.data.get("sample_type") 
        # Identificar todas as sample_names únicas dinamicamente
        
        queryset = LabTestAgriculturalSoil.objects.all()
        sample_names = queryset.values_list('sample_name', flat=True).distinct()

        # Criar a resposta HTTP para o arquivo CSV
        file_name = f"{report_name}_{sample_type}.csv"
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        # Criar o writer do CSV
        writer = csv.writer(response)

        # Escrever o cabeçalho
        header = ['substance', 'agency', 'intervention_value', 'measure'] + list(sample_names)
        writer.writerow(header)

        # Agrupar os dados por substance
        substances = queryset.values('substance', 'agency', 'intervention_value', 'measure').distinct()

        for substance_data in substances:
            # Linha base com substance, agency e intervention_value
            row = [
                substance_data['substance'],
                substance_data['agency'],
                substance_data['intervention_value'],
                substance_data['measure']
            ]

            # Adicionar os resultados para cada sample_name
            results = []
            for sample_name in sample_names:
                result = queryset.filter(
                    substance=substance_data['substance'],
                    sample_name=sample_name
                ).first()

                # Adicionar o result ou vazio se não houver dados
                if result:
                    if result.result > 0:
                        results.append(result.result)
                    else: results.append(result.under_limit)
                else: results.append('Não Testado')

            # Completar a linha e escrever no CSV
            writer.writerow(row + results)

        return response
    


class ExportLabTestAgriculturalSoilXLSX(APIView):
    """
    API para exportar dados de LabTestWater como XLSX.
    O nome do arquivo inclui report_name e sample_type.
    """

    def get(self, request):
        report_name = request.data.get("report_name")   
        sample_type = request.data.get("sample_type") 

        # Identificar todas as sample_names únicas dinamicamente
        queryset = LabTestAgriculturalSoil.objects.all()
        sample_names = queryset.values_list('sample_name', flat=True).distinct()

        # Criar o arquivo Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "LabTestWater Data"

        # Escrever o cabeçalho
        header = ['substance', 'agency', 'intervention_value', 'measure'] + list(sample_names)
        for col_num, column_title in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = column_title
            sheet[f"{col_letter}1"].font = Font(bold=True)

        # Agrupar os dados por substance
        substances = queryset.values('substance', 'agency', 'intervention_value', 'measure').distinct()

        row_number = 2  # Começa na segunda linha (primeira é o cabeçalho)
        for substance_data in substances:
            # Linha base com substance, agency e intervention_value
            row = [
                substance_data['substance'],
                substance_data['agency'],
                substance_data['intervention_value'],
                substance_data['measure']
            ]

            # Adicionar os resultados para cada sample_name
            results = []
            for sample_name in sample_names:
                result = queryset.filter(
                    substance=substance_data['substance'],
                    sample_name=sample_name
                ).first()

                if result:
                    if result.result > 0:
                        results.append(result.result)
                    else:
                        results.append(result.under_limit)
                else:
                    results.append('Não Testado')

            # Completar a linha com os dados
            full_row = row + results

            for col_num, cell_value in enumerate(full_row, 1):
                cell = sheet.cell(row=row_number, column=col_num, value=cell_value)

                # Aplicar a formatação condicional apenas para os valores de result
                if col_num > 4:  # Resultados começam na 5ª coluna
                    if isinstance(cell_value,float):    
                        intervention_value = substance_data['intervention_value']
                        if intervention_value is not None:
                            if cell_value > intervention_value:
                                cell.font = Font(color="FF0000") 
                    else:
                        cell.font = Font(color="8a8a8a")            

            row_number += 1

        # Criar a resposta HTTP para o arquivo XLSX
        file_name = f"{report_name}_{sample_type}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        workbook.save(response)
        return response        
    
######################################### Solo Residencial #######################################################################



class ExportLabTestResidentialSoilCSV(APIView):
    """
    API para exportar dados de LabTestWater como CSV.
    O nome do arquivo inclui report_name e sample_type.
    """

    def get(self, request):
        report_name = request.data.get("report_name")   
        sample_type = request.data.get("sample_type") 
        # Identificar todas as sample_names únicas dinamicamente
        
        queryset = LabTestResidentialSoil.objects.all()
        sample_names = queryset.values_list('sample_name', flat=True).distinct()

        # Criar a resposta HTTP para o arquivo CSV
        file_name = f"{report_name}_{sample_type}.csv"
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        # Criar o writer do CSV
        writer = csv.writer(response)

        # Escrever o cabeçalho
        header = ['substance', 'agency', 'intervention_value', 'measure'] + list(sample_names)
        writer.writerow(header)

        # Agrupar os dados por substance
        substances = queryset.values('substance', 'agency', 'intervention_value', 'measure').distinct()

        for substance_data in substances:
            # Linha base com substance, agency e intervention_value
            row = [
                substance_data['substance'],
                substance_data['agency'],
                substance_data['intervention_value'],
                substance_data['measure']
            ]

            # Adicionar os resultados para cada sample_name
            results = []
            for sample_name in sample_names:
                result = queryset.filter(
                    substance=substance_data['substance'],
                    sample_name=sample_name
                ).first()

                # Adicionar o result ou vazio se não houver dados
                if result:
                    if result.result > 0:
                        results.append(result.result)
                    else: results.append(result.under_limit)
                else: results.append('Não Testado')

            # Completar a linha e escrever no CSV
            writer.writerow(row + results)

        return response
    


class ExportLabTestResidentialSoilXLSX(APIView):
    """
    API para exportar dados de LabTestWater como XLSX.
    O nome do arquivo inclui report_name e sample_type.
    """

    def get(self, request):
        report_name = request.data.get("report_name")   
        sample_type = request.data.get("sample_type") 

        # Identificar todas as sample_names únicas dinamicamente
        queryset = LabTestResidentialSoil.objects.all()
        sample_names = queryset.values_list('sample_name', flat=True).distinct()

        # Criar o arquivo Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "LabTestWater Data"

        # Escrever o cabeçalho
        header = ['substance', 'agency', 'intervention_value', 'measure'] + list(sample_names)
        for col_num, column_title in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = column_title
            sheet[f"{col_letter}1"].font = Font(bold=True)

        # Agrupar os dados por substance
        substances = queryset.values('substance', 'agency', 'intervention_value', 'measure').distinct()

        row_number = 2  # Começa na segunda linha (primeira é o cabeçalho)
        for substance_data in substances:
            # Linha base com substance, agency e intervention_value
            row = [
                substance_data['substance'],
                substance_data['agency'],
                substance_data['intervention_value'],
                substance_data['measure']
            ]

            # Adicionar os resultados para cada sample_name
            results = []
            for sample_name in sample_names:
                result = queryset.filter(
                    substance=substance_data['substance'],
                    sample_name=sample_name
                ).first()

                if result:
                    if result.result > 0:
                        results.append(result.result)
                    else:
                        results.append(result.under_limit)
                else:
                    results.append('Não Testado')

            # Completar a linha com os dados
            full_row = row + results

            for col_num, cell_value in enumerate(full_row, 1):
                cell = sheet.cell(row=row_number, column=col_num, value=cell_value)

                # Aplicar a formatação condicional apenas para os valores de result
                if col_num > 4:  # Resultados começam na 5ª coluna
                    if isinstance(cell_value,float):    
                        intervention_value = substance_data['intervention_value']
                        if intervention_value is not None:
                            if cell_value > intervention_value:
                                cell.font = Font(color="FF0000") 
                    else:
                        cell.font = Font(color="8a8a8a")            

            row_number += 1

        # Criar a resposta HTTP para o arquivo XLSX
        file_name = f"{report_name}_{sample_type}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        workbook.save(response)
        return response      
    
######################################### Vapor Industrial #######################################################################

class ExportLabTestIndustrialSteamCSV(APIView):
    """
    API para exportar dados de LabTestWater como CSV.
    O nome do arquivo inclui report_name e sample_type.
    """

    def get(self, request):
        report_name = request.data.get("report_name")   
        sample_type = request.data.get("sample_type") 
        # Identificar todas as sample_names únicas dinamicamente
        
        queryset = LabTestIndustrialSteam.objects.all()
        sample_names = queryset.values_list('sample_name', flat=True).distinct()

        # Criar a resposta HTTP para o arquivo CSV
        file_name = f"{report_name}_{sample_type}.csv"
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        # Criar o writer do CSV
        writer = csv.writer(response)

        # Escrever o cabeçalho
        header = ['substance', 'agency', 'carcinogenic', 'noncarcinogenic', 'rsl' , 'measure'] + list(sample_names)
        writer.writerow(header)

        # Agrupar os dados por substance
        substances = queryset.values('substance', 'agency', 'carcinogenic', 'noncarcinogenic', 'rsl', 'measure').distinct()

        for substance_data in substances:
            # Linha base com substance, agency e intervention_value
            row = [
                substance_data['substance'],
                substance_data['agency'],
                substance_data['carcinogenic'],
                substance_data['noncarcinogenic'],
                substance_data['rsl'],
                substance_data['measure']
            ]

            # Adicionar os resultados para cada sample_name
            results = []
            for sample_name in sample_names:
                result = queryset.filter(
                    substance=substance_data['substance'],
                    sample_name=sample_name
                ).first()

                # Adicionar o result ou vazio se não houver dados
                if result:
                    if result.result > 0:
                        results.append(result.result)
                    else: results.append(result.under_limit)
                else: results.append('Não Testado')

            # Completar a linha e escrever no CSV
            writer.writerow(row + results)

        return response
    


class ExportLabTestIndustrialSteamXLSX(APIView):
    """
    API para exportar dados de LabTestWater como XLSX.
    O nome do arquivo inclui report_name e sample_type.
    """

    def get(self, request):
        report_name = request.data.get("report_name")   
        sample_type = request.data.get("sample_type") 

        # Identificar todas as sample_names únicas dinamicamente
        queryset = LabTestIndustrialSteam.objects.all()
        sample_names = queryset.values_list('sample_name', flat=True).distinct()

        # Criar o arquivo Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "LabTestWater Data"

        # Escrever o cabeçalho
        header = ['substance', 'agency', 'carcinogenic', 'noncarcinogenic', 'rsl', 'measure'] + list(sample_names)
        for col_num, column_title in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = column_title
            sheet[f"{col_letter}1"].font = Font(bold=True)

        # Agrupar os dados por substance
        substances = queryset.values('substance', 'agency', 'carcinogenic', 'noncarcinogenic', 'rsl', 'measure').distinct()

        row_number = 2  # Começa na segunda linha (primeira é o cabeçalho)
        for substance_data in substances:
            # Linha base com substance, agency e intervention_value
            row = [
                substance_data['substance'],
                substance_data['agency'],
                substance_data['carcinogenic'],
                substance_data['noncarcinogenic'],
                substance_data['rsl'],
                substance_data['measure']
            ]

            # Adicionar os resultados para cada sample_name
            results = []
            for sample_name in sample_names:
                result = queryset.filter(
                    substance=substance_data['substance'],
                    sample_name=sample_name
                ).first()

                if result:
                    if result.result > 0:
                        results.append(result.result)
                    else:
                        results.append(result.under_limit)
                else:
                    results.append('Não Testado')

            # Completar a linha com os dados
            full_row = row + results

            for col_num, cell_value in enumerate(full_row, 1):
                cell = sheet.cell(row=row_number, column=col_num, value=cell_value)

                # Aplicar a formatação condicional apenas para os valores de result
                if col_num > 4:  # Resultados começam na 5ª coluna
                    if isinstance(cell_value,float):    
                        rsl = substance_data['rsl']
                        if rsl is not None:
                            if cell_value > rsl:
                                cell.font = Font(color="FF0000") 
                    else:
                        cell.font = Font(color="8a8a8a")            

            row_number += 1

        # Criar a resposta HTTP para o arquivo XLSX
        file_name = f"{report_name}_{sample_type}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        workbook.save(response)
        return response


######################################### Vapor Residencial #######################################################################    


class ExportLabTestIndustrialSteamCSV(APIView):
    """
    API para exportar dados de LabTestWater como CSV.
    O nome do arquivo inclui report_name e sample_type.
    """

    def get(self, request):
        report_name = request.data.get("report_name")   
        sample_type = request.data.get("sample_type") 
        # Identificar todas as sample_names únicas dinamicamente
        
        queryset = LabTestIndustrialSteam.objects.all()
        sample_names = queryset.values_list('sample_name', flat=True).distinct()

        # Criar a resposta HTTP para o arquivo CSV
        file_name = f"{report_name}_{sample_type}.csv"
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        # Criar o writer do CSV
        writer = csv.writer(response)

        # Escrever o cabeçalho
        header = ['substance', 'agency', 'carcinogenic', 'noncarcinogenic', 'rsl' , 'measure'] + list(sample_names)
        writer.writerow(header)

        # Agrupar os dados por substance
        substances = queryset.values('substance', 'agency', 'carcinogenic', 'noncarcinogenic', 'rsl', 'measure').distinct()

        for substance_data in substances:
            # Linha base com substance, agency e intervention_value
            row = [
                substance_data['substance'],
                substance_data['agency'],
                substance_data['carcinogenic'],
                substance_data['noncarcinogenic'],
                substance_data['rsl'],
                substance_data['measure']
            ]

            # Adicionar os resultados para cada sample_name
            results = []
            for sample_name in sample_names:
                result = queryset.filter(
                    substance=substance_data['substance'],
                    sample_name=sample_name
                ).first()

                # Adicionar o result ou vazio se não houver dados
                if result:
                    if result.result > 0:
                        results.append(result.result)
                    else: results.append(result.under_limit)
                else: results.append('Não Testado')

            # Completar a linha e escrever no CSV
            writer.writerow(row + results)

        return response
    


class ExportLabTestResidentialSteamXLSX(APIView):
    """
    API para exportar dados de LabTestWater como XLSX.
    O nome do arquivo inclui report_name e sample_type.
    """

    def get(self, request):
        report_name = request.data.get("report_name")   
        sample_type = request.data.get("sample_type") 

        # Identificar todas as sample_names únicas dinamicamente
        queryset = LabTestResidentialSteam.objects.all()
        sample_names = queryset.values_list('sample_name', flat=True).distinct()

        # Criar o arquivo Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "LabTestWater Data"

        # Escrever o cabeçalho
        header = ['substance', 'agency', 'carcinogenic', 'noncarcinogenic', 'rsl', 'measure'] + list(sample_names)
        for col_num, column_title in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = column_title
            sheet[f"{col_letter}1"].font = Font(bold=True)

        # Agrupar os dados por substance
        substances = queryset.values('substance', 'agency', 'carcinogenic', 'noncarcinogenic', 'rsl', 'measure').distinct()

        row_number = 2  # Começa na segunda linha (primeira é o cabeçalho)
        for substance_data in substances:
            # Linha base com substance, agency e intervention_value
            row = [
                substance_data['substance'],
                substance_data['agency'],
                substance_data['carcinogenic'],
                substance_data['noncarcinogenic'],
                substance_data['rsl'],
                substance_data['measure']
            ]

            # Adicionar os resultados para cada sample_name
            results = []
            for sample_name in sample_names:
                result = queryset.filter(
                    substance=substance_data['substance'],
                    sample_name=sample_name
                ).first()

                if result:
                    if result.result > 0:
                        results.append(result.result)
                    else:
                        results.append(result.under_limit)
                else:
                    results.append('Não Testado')

            # Completar a linha com os dados
            full_row = row + results

            for col_num, cell_value in enumerate(full_row, 1):
                cell = sheet.cell(row=row_number, column=col_num, value=cell_value)

                # Aplicar a formatação condicional apenas para os valores de result
                if col_num > 4:  # Resultados começam na 5ª coluna
                    if isinstance(cell_value,float):    
                        rsl = substance_data['rsl']
                        if rsl is not None:
                            if cell_value > rsl:
                                cell.font = Font(color="FF0000") 
                    else:
                        cell.font = Font(color="8a8a8a")            

            row_number += 1

        # Criar a resposta HTTP para o arquivo XLSX
        file_name = f"{report_name}_{sample_type}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        workbook.save(response)
        return response